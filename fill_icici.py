"""
fill_icici.py - Fill ICICI brokerage Excel files from ICICI PRU source Excel files.
Source values are in decimal fraction form (e.g. 0.00692 = 0.692%); multiply by 100.
"""
import openpyxl, re, os, sys

sys.stdout.reconfigure(encoding='utf-8')

BASE = r'c:\Users\tejas.rai\Desktop\mutual fund\RE_ Brokerage Structure & Formats for the month of April to June 2026'


def norm(s):
    s = str(s).lower().strip()
    s = s.replace('icici prudential', 'icici pru')
    s = re.sub(r'\s*&\s*', ' and ', s)
    s = re.sub(r"'", ' ', s)
    s = re.sub(r'[-]', ' ', s)
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    s = s.replace('flexicap', 'flexi cap').replace('midcap', 'mid cap').replace('multicap', 'multi cap')
    s = s.replace('smallcap', 'small cap').replace('largecap', 'large cap')
    # FOF normalization
    s = s.replace('fof', 'fund of fund')
    # Name variant normalization
    s = s.replace('liquid plan', 'liquid fund')
    s = s.replace(' cb fund', ' corporate bond fund')
    s = s.replace('omni fund of fund', 'active fund of fund')  # Omni = renamed from Active
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def best_match(name, data):
    if name in data: return data[name]
    nl = name.lower()
    for k, v in data.items():
        if k.lower() == nl: return v
    nn = norm(name)
    for k, v in data.items():
        if norm(k) == nn: return v
    for k, v in data.items():
        kn = norm(k)
        if nn in kn or kn in nn: return v
    # word-subset: all words of shorter in longer, max 2 extra words
    nn_words = set(nn.split())
    for k, v in data.items():
        kn_words = set(norm(k).split())
        if len(nn_words) >= 3 and len(kn_words) >= 3:
            if nn_words.issubset(kn_words) and len(kn_words) - len(nn_words) <= 2:
                return v
            if kn_words.issubset(nn_words) and len(nn_words) - len(kn_words) <= 2:
                return v
    return None


def parse_icici_source(fname):
    """Read ICICI PRU source Excel -> {scheme_name: (t1, t2, t3, t4)}
    Source cols: B=SchemeName, C=Trail1, D=Trail2, E=Trail3, F=Trail4, G=Trail5+
    Values are decimal fractions; multiply by 100 to get percent form."""
    wb = openpyxl.load_workbook(os.path.join(BASE, fname))
    ws = wb.active
    data = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row, 2).value
        t1 = ws.cell(row, 3).value
        t2 = ws.cell(row, 4).value
        t3 = ws.cell(row, 5).value
        t4 = ws.cell(row, 6).value
        if not name or not isinstance(t1, (int, float)):
            continue
        vals = (
            round(float(t1) * 100, 4),
            round(float(t2) * 100, 4),
            round(float(t3) * 100, 4),
            round(float(t4) * 100, 4),
        )
        key = str(name).strip()
        data[key] = vals
        # Aliases for renamed or alternate-named funds
        ku = key.upper()
        if 'REGULAR GOLD SAVINGS' in ku:
            data['ICICI Prudential Gold ETF FOF'] = vals
            data['ICICI Prudential Gold ETF'] = vals
        if 'BHARAT 22' in ku:
            data['ICICI Pru Bharat 22 FOF'] = vals
    return data


def fill_icici(target_fname, source_data):
    path = os.path.join(BASE, target_fname)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdr = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    sc  = hdr.get('SchemeName', 4)
    t1c = hdr.get('Trail1stYear')
    t2c = hdr.get('Trail2ndYear')
    t3c = hdr.get('Trail3rdYear')
    t4c = hdr.get('Trail4thYearOnwards')

    filled, not_found = [], []
    for row in range(2, ws.max_row + 1):
        scheme = ws.cell(row, sc).value
        if not scheme:
            continue
        match = best_match(str(scheme), source_data)
        if match:
            t1, t2, t3, t4 = match
            if t1c: ws.cell(row, t1c).value = t1
            if t2c: ws.cell(row, t2c).value = t2
            if t3c: ws.cell(row, t3c).value = t3
            if t4c: ws.cell(row, t4c).value = t4
            filled.append(str(scheme))
        else:
            not_found.append(str(scheme))
    wb.save(path)
    return filled, not_found


pairs = [
    ("ICICI - April 2026.xlsx",  "ICICI PRU Brokerage Structure April' 26.xlsx"),
    ("ICICI -May 2026.xlsx",     "ICICI PRU Brokerage Structure May' 26.xlsx"),
    ("ICICI-June 2026.xlsx",     "ICICI PRU Brokerage Structure June' 26.xlsx"),
]

for target, source in pairs:
    print(f'\n{"="*60}')
    print(f'Processing: {target}')
    src_data = parse_icici_source(source)
    print(f'  Source schemes found: {len(src_data)}')
    filled, not_found = fill_icici(target, src_data)
    print(f'  Filled: {len(filled)} schemes')
    if not_found:
        print(f'  NOT MATCHED ({len(not_found)}):')
        for s in not_found:
            print(f'    - {s}')

print('\nDONE.')
